from openpyxl import Workbook, load_workbook

workbook = load_workbook("/home/thefireatom/Downloads/IKTST_2_k_vesna_23_24.xlsx")
ws = workbook.active
table = []

subject_col = "GI"
day_col = "FY"
week_col = "GC"
para_col = "FZ"

# index_row = []

# for i in range(1, ws.max_row):
#     if ws.cell(i, 1).value is None:
#         index_row.append(i)

# for row_del in range (len(index_row)):
#     ws.delete_rows(idx=index_row[row_del], amount=1)
#     index_row = list(map(lambda k: k - 1, index_row))

print("")
print("Расписание группы БПБО-02-22: ", "\n")
print("Нечётная неделя (I): ", "\n")

for subject, day, week, para in zip(ws[subject_col], ws[day_col], ws[week_col], ws[para_col]):

    if ((subject.value != None) and len(ws[week_col]) % 2 == 0):
        print(para.value, subject.value)

    for row in range(2, 88):
        if row % 2 == 0:
            print(para.value, subject.value)
