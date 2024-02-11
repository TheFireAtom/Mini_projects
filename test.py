from openpyxl import Workbook, load_workbook

workbook = load_workbook("/home/thefireatom/Downloads/IKTST_2_k_vesna_23_24.xlsx")
ws = workbook.active

day_col = "FY"
para_col = "FZ"
subject_col = "GI"
# week_col = "GC"

print("")
print("Расписание группы БПБО-02-22: ", "\n")
print("Нечётная неделя (I): ", "\n")

for subject, day, para in zip(ws[subject_col], ws[day_col], ws[para_col]):

    if (day.value == "ПОНЕДЕЛЬНИК"):
        print(day.value + ":", "\n")

    elif (day.value == "ВТОРНИК"):
        print(f"\n{day.value}:\n")


    elif (day.value == "СРЕДА"):
        print(f"\n{day.value}:\n")


    elif (day.value == "ЧЕТВЕРГ"):
        print(f"\n{day.value}:\n")


    elif (day.value == "ПЯТНИЦА"):
        print(f"\n{day.value}:\n")


    elif (day.value == "СУББОТА"):
        print(f"\n{day.value}:\n")
        
    if (para.value == "1"):
        print(para.value + ".")
    if (para.value == "2"):
        print(para.value + ".")
    if (para.value == "3"):
        print(para.value + ".")
    if (para.value == "4"):
        print(para.value + ".")
    if (para.value == "5"):
        print(para.value + ".")
    if (para.value == "6"):
        print(para.value + ".")
    if (para.value == "7"):
        print(para.value + ".")

    



    # start_row = 2
    # end_row = 87

    # for row_number, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row), start=start_row):
    #     if row_number % 2 == 0:
    #             print(subject.value)
    #     print()
    
